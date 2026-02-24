"""File generator tools backed by code_executor.

Provide stable `xlsx` / `pdf` capabilities for planner-executed actions.
"""

from __future__ import annotations

from typing import Any

from src.skills.base import BaseTool, ToolResult
from src.skills.code_executor import CodeExecutorTool


class XlsxGeneratorTool(BaseTool):
    @property
    def name(self) -> str:
        return "xlsx"

    @property
    def description(self) -> str:
        return (
            "Generate an XLSX spreadsheet file. "
            "If context_data contains {results:[...]}, it will convert rows to a table."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "Output XLSX filename", "default": "report.xlsx"},
                "sheet_name": {"type": "string", "description": "Worksheet name", "default": "Data"},
                "context_data": {
                    "type": "string",
                    "description": "JSON payload. Prefer {results:[{...}]} for tabular export.",
                },
            },
            "required": [],
        }

    async def execute(
        self,
        filename: str = "report.xlsx",
        sheet_name: str = "Data",
        context_data: str | None = None,
        **_: Any,
    ) -> ToolResult:
        safe_filename = (filename or "report.xlsx").strip() or "report.xlsx"
        if not safe_filename.lower().endswith(".xlsx"):
            safe_filename = f"{safe_filename}.xlsx"

        code = f"""
import json
import re
import urllib.request
from openpyxl import Workbook

data = json.load(open('context.json', encoding='utf-8'))
rows = data.get('results', [])
user_request = str(data.get('user_request', '') or '')
if not isinstance(rows, list):
    rows = []

def _to_text(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)

def _parse_numeric(token):
    token = (token or '').replace(',', '').strip()
    if not token:
        return None
    try:
        return float(token)
    except Exception:
        return None

def _extract_time_series_rows(items):
    year_value = {{}}
    year_re = re.compile(r'\\b((?:19|20)\\d{{2}})\\b')
    num_re = re.compile(r'(?<!\\d)(\\d{{1,3}}(?:,\\d{{3}})*(?:\\.\\d+)?)(?!\\d)')

    for item in items:
        if not isinstance(item, dict):
            continue
        title = str(item.get('title', '') or '')
        url = str(item.get('url', '') or '')
        content = str(item.get('content') or item.get('snippet') or '')
        text = "\\n".join([title, content])
        for line in text.splitlines():
            years = year_re.findall(line)
            if not years:
                continue
            numbers = num_re.findall(line)
            if not numbers:
                continue
            parsed_numbers = [_parse_numeric(n) for n in numbers]
            parsed_numbers = [n for n in parsed_numbers if n is not None]
            if not parsed_numbers:
                continue
            for y in years:
                yi = int(y)
                candidates = [n for n in parsed_numbers if int(n) != yi]
                if not candidates:
                    continue
                # pick the largest magnitude in the same line as representative value
                chosen = max(candidates, key=lambda x: abs(x))
                prev = year_value.get(yi)
                if prev is None or abs(chosen) > abs(prev.get('value_num', 0)):
                    year_value[yi] = {{
                        'year': yi,
                        'value': str(chosen),
                        'value_num': chosen,
                        'source_title': title,
                        'source_url': url,
                    }}

    series = [v for _, v in sorted(year_value.items(), key=lambda kv: kv[0])]
    for row in series:
        row.pop('value_num', None)
    return series

def _extract_requested_years(text):
    if not text:
        return None
    for pat in [r'近\\s*(\\d{{1,3}})\\s*年', r'过去\\s*(\\d{{1,3}})\\s*年', r'last\\s*(\\d{{1,3}})\\s*years?']:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            try:
                n = int(m.group(1))
                if n > 0:
                    return n
            except Exception:
                return None
    return None

def _has_year_value_schema(rows):
    if not rows or not isinstance(rows, list):
        return False
    first = rows[0]
    if not isinstance(first, dict):
        return False
    keys = {{str(k).lower() for k in first.keys()}}
    return ('year' in keys or '年份' in keys) and ('value' in keys or '数值' in keys or '值' in keys)

def _fetch_worldbank_series(user_text, years):
    text = (user_text or '').lower()
    indicator = None
    indicator_name = None
    if 'gdp' in text or '国内生产总值' in text:
        if any(k in text for k in ['人民币', 'rmb', 'cny', '元']):
            indicator = 'NY.GDP.MKTP.CN'
            indicator_name = 'GDP (current LCU, CNY)'
        else:
            # Default to USD for cross-country comparability and clearer scale.
            indicator = 'NY.GDP.MKTP.CD'
            indicator_name = 'GDP (current US$)'
    if not indicator:
        return []

    country = None
    if '中国' in user_text or 'china' in text:
        country = 'CHN'
    if not country:
        return []

    # World Bank API returns latest first.
    url = f'https://api.worldbank.org/v2/country/{{country}}/indicator/{{indicator}}?format=json&per_page=200'
    try:
        req = urllib.request.Request(url, headers={{'User-Agent': 'semibot-xlsx-generator'}})
        with urllib.request.urlopen(req, timeout=10) as resp:
            body = resp.read().decode('utf-8', errors='ignore')
        if not body:
            return []
        payload = json.loads(body)
        if not isinstance(payload, list) or len(payload) < 2 or not isinstance(payload[1], list):
            return []
        valid_points = []
        for item in payload[1]:
            year = int(item.get('date', 0) or 0)
            value = item.get('value')
            if not year or value in (None, ''):
                continue
            valid_points.append((year, value))
        if not valid_points:
            return []

        latest_available_year = max(y for y, _ in valid_points)
        start_year = latest_available_year - (years or 40) + 1
        out = []
        for year, value in valid_points:
            if year < start_year or year > latest_available_year:
                continue
            out.append({{
                'year': year,
                'value': str(value),
                'indicator': indicator_name,
                'unit': 'US$' if indicator.endswith('.CD') else 'CNY',
                'source_title': 'World Bank Open Data',
                'source_url': url,
            }})
        out.sort(key=lambda r: r['year'])
        # Ensure deterministic "近N年" window if extra rows appear.
        if years and len(out) > years:
            out = out[-years:]
        return out
    except Exception:
        return []

structured_rows = []
if rows and isinstance(rows[0], dict):
    # Prefer existing tabular rows if they look already structured.
    has_year_key = any(
        isinstance(r, dict) and any(str(k).lower() in ('year', '年份') for k in r.keys())
        for r in rows
    )
    if has_year_key:
        structured_rows = rows
    else:
        extracted = _extract_time_series_rows(rows)
        # Use extracted series when it is substantial; otherwise fallback to raw rows.
        if len(extracted) >= 8:
            structured_rows = extracted
        else:
            structured_rows = rows

years_requested = _extract_requested_years(user_request)
min_rows = 0
if years_requested:
    min_rows = 30 if years_requested >= 30 else max(8, years_requested // 2)
is_gdp_request = bool(
    user_request
    and (
        ('gdp' in user_request.lower())
        or ('国内生产总值' in user_request)
    )
)
prefer_worldbank = bool(
    user_request
    and years_requested
    and is_gdp_request
)
if (prefer_worldbank or not structured_rows or len(structured_rows) < min_rows) and user_request:
    wb_rows = _fetch_worldbank_series(user_request, years_requested or 40)
    if wb_rows:
        structured_rows = wb_rows

# For GDP long-range requests, never export raw link/snippet rows as table data.
# Either provide real yearly values or fail fast so planner can retry with a better path.
if is_gdp_request and years_requested:
    looks_like_year_value_table = bool(
        structured_rows
        and isinstance(structured_rows[0], dict)
        and ('year' in structured_rows[0])
        and ('value' in structured_rows[0])
    )
    if (not looks_like_year_value_table) or (len(structured_rows) < min_rows):
        raise RuntimeError(
            f"GDP time-series data unavailable: expected at least {{min_rows}} yearly rows, "
            f"got {{len(structured_rows)}}. Do not fallback to link-only rows."
        )

# Generic acceptance gate for long-range yearly table requests:
# require a real year/value table with enough rows, otherwise fail fast.
if years_requested:
    min_year_rows = min_rows or max(8, years_requested // 2)
    if (not _has_year_value_schema(structured_rows)) or (len(structured_rows) < min_year_rows):
        raise RuntimeError(
            "XLSX acceptance check failed: expected a year/value time-series table "
            f"with >= {{min_year_rows}} rows for a {{years_requested}}-year request, "
            f"but got {{len(structured_rows)}} rows."
        )

wb = Workbook()
ws = wb.active
ws.title = {sheet_name!r}

if structured_rows and isinstance(structured_rows[0], dict):
    headers = []
    seen = set()
    for r in structured_rows:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                headers.append(k)
    if not headers:
        headers = ["value"]
    ws.append(headers)
    for r in structured_rows:
        if isinstance(r, dict):
            ws.append([_to_text(r.get(h)) for h in headers])
        else:
            ws.append([_to_text(r)])
else:
    ws.append(["message"])
    ws.append(["No structured rows available in context_data.results"])

wb.save({safe_filename!r})
print("xlsx generated")
"""
        executor = CodeExecutorTool()
        return await executor.execute(
            language="python",
            code=code,
            context_data=context_data,
        )


class PdfGeneratorTool(BaseTool):
    @property
    def name(self) -> str:
        return "pdf"

    @property
    def description(self) -> str:
        return (
            "Generate a PDF file from context data. "
            "If context_data has results, they are summarized into the document."
        )

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "filename": {"type": "string", "description": "Output PDF filename", "default": "report.pdf"},
                "title": {"type": "string", "description": "PDF title", "default": "Report"},
                "context_data": {
                    "type": "string",
                    "description": "JSON payload. Prefer {results:[...]} for summary content.",
                },
            },
            "required": [],
        }

    async def execute(
        self,
        filename: str = "report.pdf",
        title: str = "Report",
        context_data: str | None = None,
        **_: Any,
    ) -> ToolResult:
        safe_filename = (filename or "report.pdf").strip() or "report.pdf"
        if not safe_filename.lower().endswith(".pdf"):
            safe_filename = f"{safe_filename}.pdf"

        code = f"""
import json
import re
from fpdf import FPDF

data = json.load(open('context.json', encoding='utf-8'))
rows = data.get('results', [])
user_request = str(data.get('user_request', '') or '')
generated_at = str(data.get('generated_at', '') or '')
if not isinstance(rows, list):
    rows = []

pdf = FPDF()
pdf.add_page()
pdf.set_margins(10, 10, 10)
pdf.set_auto_page_break(auto=True, margin=15)
font_path = '/System/Library/Fonts/Hiragino Sans GB.ttc'
use_cjk = False
try:
    pdf.add_font('HiraginoGB', '', font_path)
    pdf.set_font('HiraginoGB', size=16)
    use_cjk = True
except Exception:
    pdf.set_font('Helvetica', size=16)

pdf.cell(0, 10, text={title!r}, new_x='LMARGIN', new_y='NEXT')

if use_cjk:
    pdf.set_font('HiraginoGB', size=11)
else:
    pdf.set_font('Helvetica', size=11)

if user_request:
    pdf.multi_cell(0, 7, text='User request: ' + user_request[:300])
if generated_at:
    pdf.multi_cell(0, 7, text='Generated at: ' + generated_at[:40])
pdf.ln(2)

def _norm(s):
    return str(s or '').replace('\\n', ' ').strip()

def _is_placeholder_text(s):
    t = _norm(s)
    if not t:
        return True
    if t in {'...', '…', '-', '--', 'n/a', 'N/A'}:
        return True
    if len(t) <= 6 and set(t) <= set('.-_~ '):
        return True
    return False

def _is_stock_research_request(text):
    t = (text or '').lower()
    keys = ['股票', '股价', '港股', '美股', 'a股', 'target price', 'ticker', 'stock', 'equity', '估值', '财报']
    return any(k in t for k in keys)

def _extract_evidence_points(items, limit=12):
    points = []
    # Keep only lines that look like factual evidence (numbers + finance keywords).
    kw = re.compile(r'(营收|收入|利润|净利|eps|市盈率|估值|目标价|评级|回购|分红|指引|同比|环比|guidance|growth|revenue|margin|earnings|target|valuation|pe|ebitda)', re.IGNORECASE)
    num = re.compile(r'(\\d[\\d,\\.]*\\s*(%|亿|万|hk\\$|usd|rmb|港元|美元|人民币)?)', re.IGNORECASE)
    for i, row in enumerate(items):
        if not isinstance(row, dict):
            continue
        title = _norm(row.get('title') or row.get('name') or f'Result {{i+1}}')
        content = _norm(row.get('content') or row.get('snippet'))
        if not content or _is_placeholder_text(content):
            continue
        for seg in re.split(r'[。；;\\n]', content):
            s = _norm(seg)
            if len(s) < 12:
                continue
            has_num = bool(num.search(s))
            has_kw = bool(kw.search(s))
            has_year = bool(re.search(r'\\b(19|20)\\d{{2}}\\b', s))
            if has_num and (has_kw or has_year):
                points.append((title, s[:180]))
                if len(points) >= limit:
                    return points
    return points

def _extract_summary_points(items, limit=5):
    points = []
    seen = set()
    for i, row in enumerate(items):
        if not isinstance(row, dict):
            continue
        title = _norm(row.get('title') or row.get('name') or '')
        content = _norm(row.get('content') or row.get('snippet'))
        candidates = []
        if title and len(title) >= 4 and not _is_placeholder_text(title):
            candidates.append(title)
        if content:
            for seg in re.split(r'[。；;\\n]', content):
                s = _norm(seg)
                if len(s) < 10:
                    continue
                if s.lower().startswith('http'):
                    continue
                if s.count('|') > 5:
                    continue
                if _is_placeholder_text(s):
                    continue
                candidates.append(s)
                if len(candidates) >= 3:
                    break
        for c in candidates:
            key = c[:80].lower()
            if key in seen:
                continue
            seen.add(key)
            points.append(c[:180])
            if len(points) >= limit:
                return points
    return points

def _section_title(text):
    if use_cjk:
        pdf.set_font('HiraginoGB', size=13)
    else:
        pdf.set_font('Helvetica', size=13)
    pdf.multi_cell(0, 8, text=text)
    if use_cjk:
        pdf.set_font('HiraginoGB', size=11)
    else:
        pdf.set_font('Helvetica', size=11)

if rows:
    structured = [r for r in rows if isinstance(r, dict)]
    source_rows = []
    for r in structured:
        title_text = _norm(r.get('title') or r.get('name') or '')
        url_text = _norm(r.get('url'))
        content_text = _norm(r.get('content') or r.get('snippet'))
        if _is_placeholder_text(title_text) and _is_placeholder_text(content_text):
            continue
        if not title_text and not url_text and not content_text:
            continue
        source_rows.append(r)
        if len(source_rows) >= 10:
            break

    _section_title('报告范围与方法')
    pdf.multi_cell(0, 7, text='本报告基于检索到的公开网页信息自动整理，重点输出可追溯证据。以下内容不构成投资建议。')
    pdf.ln(1)

    _section_title('执行摘要')
    summary_points = _extract_summary_points(structured, limit=5)
    if summary_points:
        for p in summary_points:
            pdf.multi_cell(0, 7, text='- ' + p)
    else:
        # Use source titles as a weak-but-useful fallback to avoid empty template output.
        fallback_titles = []
        for row in source_rows:
            t = _norm(row.get('title') or row.get('name') or '')
            if t and not _is_placeholder_text(t):
                fallback_titles.append(t[:120])
            if len(fallback_titles) >= 4:
                break
        if fallback_titles:
            for t in fallback_titles:
                pdf.multi_cell(0, 7, text='- 来源显示重点：' + t)
        else:
            pdf.multi_cell(0, 7, text='本次检索命中结果有限，未提取到高质量可归纳摘要。建议补充公司公告/交易所披露后再生成完整版研究结论。')
    pdf.ln(1)

    _section_title('检索证据摘要')
    for i, row in enumerate(source_rows, start=1):
        title_text = _norm(row.get('title') or row.get('name') or f'Source {{i}}')
        url = _norm(row.get('url'))
        content = _norm(row.get('content') or row.get('snippet'))
        if len(content) > 280:
            content = content[:280] + '...'
        pdf.multi_cell(0, 7, text=f'{{i}}. ' + title_text[:150])
        if content:
            pdf.multi_cell(0, 7, text=content)
        if url:
            shown = url[:110]
            pdf.cell(0, 7, text=shown, link=url, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(1)

    evidence = _extract_evidence_points(structured)
    _section_title('关键发现（仅基于检索证据）')
    if evidence:
        for i, (src_title, point) in enumerate(evidence, start=1):
            pdf.multi_cell(0, 7, text=f'- 发现{{i}}（{{src_title[:36]}}）：{{point}}')
    else:
        pdf.multi_cell(0, 7, text='未能从检索结果中提取到足够的结构化财务数值证据，建议补充公司公告/交易所披露后再做定量结论。')
    pdf.ln(1)

    _section_title('风险与不确定性')
    pdf.multi_cell(0, 7, text='1) 数据来源可能存在转载和时效差异；2) 部分站点为观点类内容，需交叉验证；3) 对目标价/预测应以最新财报和公告为准。')
    pdf.ln(1)

    _section_title('数据可得性声明')
    pdf.multi_cell(0, 7, text='当前检索结果主要来自公开网页快照，未直接接入交易所完整披露表格或公司IR原始文件。若用于投资决策，请以最新公告与财报原文复核。')
    pdf.ln(1)

    _section_title('参考来源')
    shown = set()
    for row in source_rows:
        url = _norm(row.get('url'))
        if not url or url in shown:
            continue
        shown.add(url)
        pdf.cell(0, 7, text=url[:120], link=url, new_x='LMARGIN', new_y='NEXT')
else:
    _section_title('执行摘要')
    pdf.multi_cell(0, 7, text='当前未检索到可用于总结的结构化内容，报告仅包含任务说明。建议先执行检索步骤后再生成报告。')
    pdf.ln(1)
    _section_title('参考来源')
    pdf.multi_cell(0, 7, text='暂无来源')

pdf.output({safe_filename!r})
print("pdf generated")
print("report_summary_sections=报告范围与方法|执行摘要|检索证据摘要|关键发现（仅基于检索证据）|风险与不确定性|数据可得性声明|参考来源")
print("report_summary_source_count=" + str(len(rows) if isinstance(rows, list) else 0))
"""
        executor = CodeExecutorTool()
        return await executor.execute(
            language="python",
            code=code,
            context_data=context_data,
        )
