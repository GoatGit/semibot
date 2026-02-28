"""Builtin CSV/Excel processing tool."""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from src.server.config_store import RuntimeConfigStore
from src.skills.base import BaseTool, ToolResult
from src.storage.file_manager import FileManager

_file_manager = FileManager()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _normalize_rows(data: Any) -> list[dict[str, Any]]:
    if not isinstance(data, list):
        return []
    rows: list[dict[str, Any]] = []
    for item in data:
        if isinstance(item, dict):
            rows.append(dict(item))
        else:
            rows.append({"value": item})
    return rows


class CsvXlsxTool(BaseTool):
    """Read/write/filter/aggregate/pivot tabular data in CSV/XLSX files."""

    def __init__(self) -> None:
        self.root = Path(os.getenv("SEMIBOT_CSV_XLSX_ROOT", str(Path.home()))).resolve()
        self.max_return_rows = int(os.getenv("SEMIBOT_CSV_XLSX_MAX_RETURN_ROWS", "500"))
        self.default_sheet_name = str(os.getenv("SEMIBOT_CSV_XLSX_DEFAULT_SHEET", "Data")).strip() or "Data"
        self._load_runtime_config()

    @property
    def name(self) -> str:
        return "csv_xlsx"

    @property
    def description(self) -> str:
        return "Read/write/filter/aggregate/pivot CSV and XLSX datasets."

    @property
    def parameters(self) -> dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "action": {
                    "type": "string",
                    "enum": ["read", "write", "filter", "aggregate", "pivot"],
                    "description": "Operation to perform.",
                },
                "path": {"type": "string", "description": "Input file path under configured root."},
                "output_path": {"type": "string", "description": "Output file path for write/export operations."},
                "format": {"type": "string", "enum": ["csv", "xlsx"], "description": "Output format override."},
                "sheet_name": {"type": "string", "description": "Worksheet name for XLSX read/write."},
                "data": {"description": "Rows used for write/transform operations."},
                "filters": {
                    "type": "array",
                    "description": "Filter rules: [{field, op, value}] where op in eq/ne/gt/gte/lt/lte/contains/in.",
                },
                "group_by": {"type": "array", "description": "Fields used for aggregate grouping."},
                "metrics": {
                    "type": "array",
                    "description": "Aggregate metrics [{field, op, as}] op in count/sum/avg/min/max.",
                },
                "pivot": {
                    "type": "object",
                    "description": "Pivot config: {index, columns, values, agg}.",
                },
                "limit": {"type": "integer", "description": "Maximum rows returned in tool output."},
            },
            "required": ["action"],
        }

    def _load_runtime_config(self) -> None:
        try:
            store = RuntimeConfigStore(db_path=os.getenv("SEMIBOT_EVENTS_DB_PATH"))
            item = store.get_tool_by_name(self.name)
            config = item.get("config") if isinstance(item, dict) else {}
            if not isinstance(config, dict):
                return

            root_from_cfg = config.get("rootPath") or config.get("root")
            if isinstance(root_from_cfg, str) and root_from_cfg.strip():
                self.root = Path(root_from_cfg).expanduser().resolve()

            max_rows = config.get("maxReturnRows")
            if isinstance(max_rows, int) and max_rows > 0:
                self.max_return_rows = max_rows

            sheet = config.get("sheetName")
            if isinstance(sheet, str) and sheet.strip():
                self.default_sheet_name = sheet.strip()
        except Exception:
            return

    def _resolve_path(self, raw_path: str) -> Path:
        rel = (raw_path or "").strip()
        if not rel:
            raise ValueError("path is required")
        target = (self.root / rel).resolve()
        if target != self.root and self.root not in target.parents:
            raise ValueError(f"path escapes configured root: {self.root}")
        return target

    def _resolve_format(self, path: Path, requested: str | None) -> str:
        if requested in {"csv", "xlsx"}:
            return requested
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return "csv"
        if suffix == ".xlsx":
            return "xlsx"
        return "csv"

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            return [dict(row) for row in reader]

    def _read_xlsx(self, path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(item).strip() if item is not None else "" for item in rows[0]]
            normalized_headers = [
                header if header else f"column_{idx + 1}"
                for idx, header in enumerate(headers)
            ]
            out: list[dict[str, Any]] = []
            for row in rows[1:]:
                record: dict[str, Any] = {}
                for idx, header in enumerate(normalized_headers):
                    value = row[idx] if idx < len(row) else None
                    record[header] = value
                out.append(record)
            return out
        finally:
            wb.close()

    def _read_rows(self, path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
        if not path.exists() or not path.is_file():
            raise ValueError(f"File not found: {path}")
        file_format = self._resolve_format(path, None)
        if file_format == "xlsx":
            return self._read_xlsx(path, sheet_name)
        return self._read_csv(path)

    def _write_csv(self, path: Path, rows: list[dict[str, Any]]) -> None:
        headers: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in seen:
                    seen.add(key_text)
                    headers.append(key_text)
        if not headers:
            headers = ["value"]
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow({header: row.get(header) for header in headers})

    def _write_xlsx(self, path: Path, rows: list[dict[str, Any]], sheet_name: str | None) -> None:
        headers: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in seen:
                    seen.add(key_text)
                    headers.append(key_text)
        if not headers:
            headers = ["value"]

        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or self.default_sheet_name)[:31]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header) for header in headers])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    def _write_rows(self, path: Path, rows: list[dict[str, Any]], sheet_name: str | None, fmt: str | None) -> None:
        file_format = self._resolve_format(path, fmt)
        if file_format == "xlsx":
            final_path = path if path.suffix.lower() == ".xlsx" else Path(f"{path}.xlsx")
            self._write_xlsx(final_path, rows, sheet_name)
            return
        final_path = path if path.suffix.lower() == ".csv" else Path(f"{path}.csv")
        self._write_csv(final_path, rows)

    def _apply_filters(self, rows: list[dict[str, Any]], filters: list[dict[str, Any]]) -> list[dict[str, Any]]:
        if not filters:
            return rows

        def _match(row: dict[str, Any], rule: dict[str, Any]) -> bool:
            field = str(rule.get("field") or "").strip()
            op = str(rule.get("op") or "eq").strip().lower()
            expected = rule.get("value")
            if not field:
                return True
            actual = row.get(field)
            if op == "eq":
                return actual == expected
            if op == "ne":
                return actual != expected
            if op in {"gt", "gte", "lt", "lte"}:
                a = _to_float(actual)
                b = _to_float(expected)
                if a is None or b is None:
                    return False
                if op == "gt":
                    return a > b
                if op == "gte":
                    return a >= b
                if op == "lt":
                    return a < b
                return a <= b
            if op == "contains":
                return str(expected) in str(actual)
            if op == "in":
                if isinstance(expected, list):
                    return actual in expected
                return str(actual) in {item.strip() for item in str(expected).split(",")}
            return False

        out: list[dict[str, Any]] = []
        for row in rows:
            if all(_match(row, rule) for rule in filters):
                out.append(row)
        return out

    def _aggregate_rows(
        self,
        rows: list[dict[str, Any]],
        group_by: list[str],
        metrics: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
        for row in rows:
            key = tuple(row.get(col) for col in group_by)
            grouped.setdefault(key, []).append(row)

        metric_defs = metrics or [{"field": "*", "op": "count", "as": "count"}]
        out: list[dict[str, Any]] = []
        for key, members in grouped.items():
            item: dict[str, Any] = {}
            for idx, col in enumerate(group_by):
                item[col] = key[idx] if idx < len(key) else None

            for metric in metric_defs:
                field = str(metric.get("field") or "*").strip()
                op = str(metric.get("op") or "count").strip().lower()
                alias = str(metric.get("as") or f"{op}_{field}").strip()
                if not alias:
                    continue
                if op == "count":
                    item[alias] = len(members)
                    continue

                values = [_to_float(row.get(field)) for row in members if field in row]
                values = [value for value in values if value is not None]
                if op == "sum":
                    item[alias] = float(sum(values)) if values else 0.0
                elif op == "avg":
                    item[alias] = float(sum(values) / len(values)) if values else None
                elif op == "min":
                    item[alias] = min(values) if values else None
                elif op == "max":
                    item[alias] = max(values) if values else None
                else:
                    item[alias] = None
            out.append(item)
        return out

    def _pivot_rows(self, rows: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
        index_col = str(config.get("index") or "").strip()
        columns_col = str(config.get("columns") or "").strip()
        values_col = str(config.get("values") or "").strip()
        agg = str(config.get("agg") or "sum").strip().lower()
        if not index_col or not columns_col or not values_col:
            raise ValueError("pivot requires index, columns, and values")

        table: dict[Any, dict[Any, list[float]]] = {}
        for row in rows:
            idx = row.get(index_col)
            col = row.get(columns_col)
            val = _to_float(row.get(values_col))
            if idx is None or col is None or val is None:
                continue
            table.setdefault(idx, {}).setdefault(col, []).append(val)

        all_columns = sorted({col for value in table.values() for col in value.keys()}, key=str)
        out: list[dict[str, Any]] = []
        for idx, bucket in table.items():
            row_out: dict[str, Any] = {index_col: idx}
            for col in all_columns:
                values = bucket.get(col, [])
                if not values:
                    row_out[str(col)] = None
                    continue
                if agg == "avg":
                    row_out[str(col)] = float(sum(values) / len(values))
                elif agg == "min":
                    row_out[str(col)] = min(values)
                elif agg == "max":
                    row_out[str(col)] = max(values)
                elif agg == "count":
                    row_out[str(col)] = len(values)
                else:
                    row_out[str(col)] = float(sum(values))
            out.append(row_out)
        return out

    def _materialize_rows(self, path: str | None, sheet_name: str | None, data: Any) -> list[dict[str, Any]]:
        normalized_rows = _normalize_rows(data)
        if normalized_rows:
            return normalized_rows
        if not path:
            return []
        source_path = self._resolve_path(path)
        return self._read_rows(source_path, sheet_name)

    def _limit_rows(self, rows: list[dict[str, Any]], limit: int) -> tuple[list[dict[str, Any]], bool]:
        truncated = len(rows) > limit
        if truncated:
            return rows[:limit], True
        return rows, False

    def _persist_generated(self, path: Path) -> list[dict[str, Any]]:
        meta = _file_manager.persist_file(path)
        return [meta] if meta else []

    async def execute(
        self,
        action: str,
        path: str | None = None,
        output_path: str | None = None,
        format: str | None = None,
        sheet_name: str | None = None,
        data: Any = None,
        filters: list[dict[str, Any]] | None = None,
        group_by: list[str] | None = None,
        metrics: list[dict[str, Any]] | None = None,
        pivot: dict[str, Any] | None = None,
        limit: int | None = None,
        **_: Any,
    ) -> ToolResult:
        self._load_runtime_config()
        action_name = str(action or "").strip().lower()
        if action_name not in {"read", "write", "filter", "aggregate", "pivot"}:
            return ToolResult.error_result("action must be read/write/filter/aggregate/pivot")

        output_limit = (
            limit
            if isinstance(limit, int) and limit > 0
            else self.max_return_rows
        )

        try:
            if action_name == "read":
                if not path:
                    return ToolResult.error_result("path is required for action=read")
                source_path = self._resolve_path(path)
                rows = self._read_rows(source_path, sheet_name)
                preview, truncated = self._limit_rows(rows, output_limit)
                columns = sorted({key for row in rows for key in row.keys()})
                return ToolResult.success_result(
                    {
                        "action": "read",
                        "path": str(source_path),
                        "rows": preview,
                        "total_rows": len(rows),
                        "columns": columns,
                        "truncated": truncated,
                    }
                )

            if action_name == "write":
                rows = _normalize_rows(data)
                if not rows:
                    return ToolResult.error_result("data must be a non-empty list for action=write")
                target_raw = (output_path or path or "").strip()
                if not target_raw:
                    return ToolResult.error_result("output_path (or path) is required for action=write")
                target_path = self._resolve_path(target_raw)
                output_fmt = format if format in {"csv", "xlsx"} else self._resolve_format(target_path, format)
                if output_fmt == "xlsx" and target_path.suffix.lower() != ".xlsx":
                    target_path = Path(f"{target_path}.xlsx")
                if output_fmt == "csv" and target_path.suffix.lower() != ".csv":
                    target_path = Path(f"{target_path}.csv")
                self._write_rows(target_path, rows, sheet_name, output_fmt)
                generated_files = self._persist_generated(target_path)
                return ToolResult.success_result(
                    {
                        "action": "write",
                        "path": str(target_path),
                        "rows_written": len(rows),
                    },
                    generated_files=generated_files,
                )

            materialized = self._materialize_rows(path, sheet_name, data)
            if not materialized:
                return ToolResult.error_result("No input rows. Provide path or data.")

            result_rows: list[dict[str, Any]]
            if action_name == "filter":
                result_rows = self._apply_filters(materialized, filters or [])
            elif action_name == "aggregate":
                result_rows = self._aggregate_rows(materialized, [str(x) for x in (group_by or [])], metrics or [])
            else:
                result_rows = self._pivot_rows(materialized, pivot or {})

            generated_files: list[dict[str, Any]] = []
            if output_path and output_path.strip():
                target_path = self._resolve_path(output_path.strip())
                output_fmt = format if format in {"csv", "xlsx"} else self._resolve_format(target_path, format)
                if output_fmt == "xlsx" and target_path.suffix.lower() != ".xlsx":
                    target_path = Path(f"{target_path}.xlsx")
                if output_fmt == "csv" and target_path.suffix.lower() != ".csv":
                    target_path = Path(f"{target_path}.csv")
                self._write_rows(target_path, result_rows, sheet_name, output_fmt)
                generated_files = self._persist_generated(target_path)

            preview, truncated = self._limit_rows(result_rows, output_limit)
            return ToolResult.success_result(
                {
                    "action": action_name,
                    "rows": preview,
                    "total_rows": len(result_rows),
                    "truncated": truncated,
                    "output_path": output_path,
                },
                generated_files=generated_files,
            )
        except Exception as exc:
            return ToolResult.error_result(str(exc))
